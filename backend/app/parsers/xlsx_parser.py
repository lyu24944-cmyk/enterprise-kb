from pathlib import Path

from openpyxl import load_workbook

from app.parsers.base import ParsedBlock, ParsedDocument
from app.parsers.pdf_parser import _guess_category


def parse_xlsx(path: Path) -> ParsedDocument:
    wb = load_workbook(path, data_only=True)
    blocks: list[ParsedBlock] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[str] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if any(cells):
                rows.append(" | ".join(cells))
        if rows:
            text = f"## Sheet: {sheet_name}\n" + "\n".join(rows)
            blocks.append(
                ParsedBlock(
                    text=text,
                    section=sheet_name,
                    metadata={"type": "spreadsheet", "sheet": sheet_name},
                )
            )
    wb.close()
    return ParsedDocument(blocks=blocks, doc_type="xlsx", category=_guess_category(path.name))
